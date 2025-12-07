import pandas as pd
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from datetime import datetime, timedelta
import os
import argparse
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import traceback

class AllHandsReportGenerator:
    def __init__(self, credentials_path: str):
        """Initialize GSC and GA4 API connections"""
        self.gsc_service = None
        self.ga_service = None
        self.setup_apis(credentials_path)
    
    def setup_apis(self, credentials_path: str):
        """Setup Google Search Console and GA4 APIs"""
        try:
            credentials = Credentials.from_service_account_file(
                credentials_path, 
                scopes=[
                    'https://www.googleapis.com/auth/webmasters.readonly',
                    'https://www.googleapis.com/auth/analytics.readonly'
                ]
            )
            
            self.gsc_service = build('searchconsole', 'v1', credentials=credentials)
            self.ga_service = build('analyticsdata', 'v1beta', credentials=credentials)
            print("✓ APIs connected successfully")
            
        except Exception as e:
            print(f"❌ API setup failed: {e}")
            raise
    
    def calculate_percentage_change(self, old_value: float, new_value: float) -> float:
        """Calculate percentage change between two values"""
        if old_value == 0:
            return 100.0 if new_value > 0 else 0.0
        return round(((new_value - old_value) / old_value) * 100, 2)
    
    def get_date_ranges(self, start_date: str, end_date: str):
        """Calculate all required date ranges from start and end dates"""
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        
        # Calculate the number of days in the current period
        period_days = (end_dt - start_dt).days + 1
        
        # Current period (user-specified)
        current_start = start_dt
        current_end = end_dt
        
        # Previous period (same number of days, immediately before current)
        prev_end = start_dt - timedelta(days=1)
        prev_start = prev_end - timedelta(days=period_days - 1)
        
        # Period -2 (same number of days, before previous)
        prev2_end = prev_start - timedelta(days=1)
        prev2_start = prev2_end - timedelta(days=period_days - 1)
        
        # Year over year (same dates, 1 year ago)
        yoy_start = start_dt - timedelta(days=365)
        yoy_end = end_dt - timedelta(days=365)
        
        return {
            'current': (current_start.strftime('%Y-%m-%d'), current_end.strftime('%Y-%m-%d')),
            'previous': (prev_start.strftime('%Y-%m-%d'), prev_end.strftime('%Y-%m-%d')),
            'previous2': (prev2_start.strftime('%Y-%m-%d'), prev2_end.strftime('%Y-%m-%d')),
            'yoy': (yoy_start.strftime('%Y-%m-%d'), yoy_end.strftime('%Y-%m-%d')),
            'period_days': period_days
        }
    
    def get_gsc_weekly_data(self, site_url: str, start_date: str, end_date: str) -> pd.DataFrame:
        """Get GSC data broken down by week"""
        print(f"📊 Fetching GSC weekly data ({start_date} to {end_date})...")
        
        try:
            property_urls = [
                site_url,
                f"sc-domain:{site_url.replace('https://', '').replace('http://', '').replace('www.', '').rstrip('/')}"
            ]
            
            for property_url in property_urls:
                try:
                    request = {
                        'startDate': start_date,
                        'endDate': end_date,
                        'dimensions': ['date'],
                        'rowLimit': 25000
                    }
                    
                    response = self.gsc_service.searchanalytics().query(
                        siteUrl=property_url, 
                        body=request
                    ).execute()
                    
                    print(f"✓ Using GSC property: {property_url}")
                    break
                    
                except Exception as e:
                    continue
            else:
                print("❌ No valid GSC property found")
                return pd.DataFrame()
            
            # Process daily data into weekly
            daily_data = []
            for row in response.get('rows', []):
                daily_data.append({
                    'date': row['keys'][0],
                    'clicks': row['clicks'],
                    'impressions': row['impressions'],
                    'ctr': round(row['ctr'] * 100, 2)
                })
            
            df = pd.DataFrame(daily_data)
            if df.empty:
                return df
            
            df['date'] = pd.to_datetime(df['date'])
            df['week'] = df['date'].dt.isocalendar().week
            df['year'] = df['date'].dt.year
            df['week_start'] = df['date'] - pd.to_timedelta(df['date'].dt.dayofweek, unit='d')
            
            # Group by week
            weekly_data = df.groupby(['year', 'week', 'week_start']).agg({
                'clicks': 'sum',
                'impressions': 'sum',
                'ctr': 'mean'
            }).reset_index()
            
            weekly_data['ctr'] = weekly_data['ctr'].round(2)
            weekly_data = weekly_data.sort_values('week_start')
            
            # Calculate week-over-week changes
            weekly_data['clicks_wow_change'] = weekly_data['clicks'].pct_change() * 100
            weekly_data['impressions_wow_change'] = weekly_data['impressions'].pct_change() * 100
            weekly_data['ctr_wow_change'] = weekly_data['ctr'].pct_change() * 100
            
            # Round percentage changes
            for col in ['clicks_wow_change', 'impressions_wow_change', 'ctr_wow_change']:
                weekly_data[col] = weekly_data[col].round(2)
            
            return weekly_data
            
        except Exception as e:
            print(f"❌ GSC weekly data retrieval failed: {e}")
            return pd.DataFrame()
    
    def get_ga4_weekly_data(self, property_id: str, start_date: str, end_date: str) -> pd.DataFrame:
        """Get GA4 sessions data broken down by week"""
        print(f"📈 Fetching GA4 weekly data ({start_date} to {end_date})...")
        
        try:
            # Get daily data by channel
            request = {
                'dateRanges': [{'startDate': start_date, 'endDate': end_date}],
                'dimensions': [
                    {'name': 'date'},
                    {'name': 'sessionDefaultChannelGroup'}
                ],
                'metrics': [{'name': 'sessions'}],
                'limit': 100000
            }
            
            response = self.ga_service.properties().runReport(
                property=f'properties/{property_id}',
                body=request
            ).execute()
            
            # Process daily data
            daily_data = []
            for row in response.get('rows', []):
                date_str = row['dimensionValues'][0]['value']
                channel = row['dimensionValues'][1]['value'].lower()
                sessions = int(row['metricValues'][0]['value'])
                
                daily_data.append({
                    'date': date_str,
                    'channel': channel,
                    'sessions': sessions
                })
            
            if not daily_data:
                return pd.DataFrame()
            
            # Convert to DataFrame and pivot
            df = pd.DataFrame(daily_data)
            df['date'] = pd.to_datetime(df['date'])
            
            # Categorize channels
            def categorize_channel(channel):
                if 'organic' in channel:
                    return 'organic_sessions'
                elif any(x in channel for x in ['paid', 'display', 'video']):
                    return 'paid_sessions'
                elif 'direct' in channel:
                    return 'direct_sessions'
                else:
                    return 'other_sessions'
            
            df['channel_category'] = df['channel'].apply(categorize_channel)
            
            # Group by date and channel category
            daily_summary = df.groupby(['date', 'channel_category'])['sessions'].sum().unstack(fill_value=0)
            daily_summary['total_sessions'] = daily_summary.sum(axis=1)
            
            # Ensure all columns exist
            for col in ['organic_sessions', 'paid_sessions', 'direct_sessions', 'other_sessions']:
                if col not in daily_summary.columns:
                    daily_summary[col] = 0
            
            # Add week information
            daily_summary = daily_summary.reset_index()
            daily_summary['week'] = daily_summary['date'].dt.isocalendar().week
            daily_summary['year'] = daily_summary['date'].dt.year
            daily_summary['week_start'] = daily_summary['date'] - pd.to_timedelta(daily_summary['date'].dt.dayofweek, unit='d')
            
            # Group by week
            weekly_data = daily_summary.groupby(['year', 'week', 'week_start']).agg({
                'organic_sessions': 'sum',
                'paid_sessions': 'sum',
                'direct_sessions': 'sum',
                'other_sessions': 'sum',
                'total_sessions': 'sum'
            }).reset_index()
            
            weekly_data = weekly_data.sort_values('week_start')
            
            # Calculate week-over-week changes
            for metric in ['organic_sessions', 'paid_sessions', 'direct_sessions', 'total_sessions']:
                weekly_data[f'{metric}_wow_change'] = weekly_data[metric].pct_change() * 100
                weekly_data[f'{metric}_wow_change'] = weekly_data[f'{metric}_wow_change'].round(2)
            
            return weekly_data
            
        except Exception as e:
            print(f"❌ GA4 weekly data retrieval failed: {e}")
            return pd.DataFrame()
    
    def get_ga4_monthly_data(self, property_id: str, start_date: str, end_date: str) -> dict:
        """Get GA4 sessions data by channel for a month (aggregated)"""
        print(f"📈 Fetching GA4 monthly data ({start_date} to {end_date})...")
        
        try:
            request = {
                'dateRanges': [{'startDate': start_date, 'endDate': end_date}],
                'dimensions': [{'name': 'sessionDefaultChannelGroup'}],
                'metrics': [{'name': 'sessions'}],
                'limit': 100000
            }
            
            response = self.ga_service.properties().runReport(
                property=f'properties/{property_id}',
                body=request
            ).execute()
            
            data = {
                'organic_sessions': 0,
                'paid_sessions': 0,
                'direct_sessions': 0,
                'total_sessions': 0
            }
            
            for row in response.get('rows', []):
                channel = row['dimensionValues'][0]['value'].lower()
                sessions = int(row['metricValues'][0]['value'])
                
                data['total_sessions'] += sessions
                
                if 'organic' in channel:
                    data['organic_sessions'] += sessions
                elif any(x in channel for x in ['paid', 'display', 'video']):
                    data['paid_sessions'] += sessions
                elif 'direct' in channel:
                    data['direct_sessions'] += sessions
            
            return data
            
        except Exception as e:
            print(f"❌ GA4 monthly data retrieval failed: {e}")
            return {'organic_sessions': 0, 'paid_sessions': 0, 'direct_sessions': 0, 'total_sessions': 0}
    
    def get_gsc_url_performance(self, site_url: str, start_date: str, end_date: str) -> pd.DataFrame:
        """Get GSC data by URL for performance analysis"""
        print(f"📊 Fetching GSC URL performance ({start_date} to {end_date})...")
        
        try:
            property_urls = [
                site_url,
                f"sc-domain:{site_url.replace('https://', '').replace('http://', '').replace('www.', '').rstrip('/')}"
            ]
            
            for property_url in property_urls:
                try:
                    all_rows = []
                    start_row = 0
                    
                    while True:
                        request = {
                            'startDate': start_date,
                            'endDate': end_date,
                            'dimensions': ['page'],
                            'rowLimit': 25000,
                            'startRow': start_row
                        }
                        
                        response = self.gsc_service.searchanalytics().query(
                            siteUrl=property_url, 
                            body=request
                        ).execute()
                        
                        rows = response.get('rows', [])
                        if not rows:
                            break
                            
                        all_rows.extend(rows)
                        
                        if len(rows) < 25000:
                            break
                            
                        start_row += 25000
                        
                        if start_row >= 100000:
                            break
                    
                    break
                    
                except Exception as e:
                    continue
            else:
                return pd.DataFrame()
            
            # Process the data
            url_data = []
            for row in all_rows:
                page_url = row['keys'][0]
                
                # Minimal URL normalization - only for completely empty/root paths
                if page_url == '' or page_url == '/':
                    page_url = site_url.rstrip('/') + '/'
                elif not page_url.startswith('http'):
                    # Keep the relative path as-is, just add the domain
                    page_url = site_url.rstrip('/') + ('/' + page_url.lstrip('/') if not page_url.startswith('/') else page_url)
                
                url_data.append({
                    'url': page_url,
                    'clicks': row['clicks'],
                    'impressions': row['impressions'],
                    'ctr': round(row['ctr'] * 100, 2),
                    'position': round(row['position'], 1)
                })
            
            return pd.DataFrame(url_data)
            
        except Exception as e:
            print(f"❌ GSC URL performance retrieval failed: {e}")
            return pd.DataFrame()
    
    def get_gsc_query_performance(self, site_url: str, start_date: str, end_date: str) -> pd.DataFrame:
        """Get GSC data by query for performance analysis"""
        print(f"📊 Fetching GSC query performance ({start_date} to {end_date})...")
        
        try:
            property_urls = [
                site_url,
                f"sc-domain:{site_url.replace('https://', '').replace('http://', '').replace('www.', '').rstrip('/')}"
            ]
            
            for property_url in property_urls:
                try:
                    all_rows = []
                    start_row = 0
                    
                    while True:
                        request = {
                            'startDate': start_date,
                            'endDate': end_date,
                            'dimensions': ['query'],
                            'rowLimit': 25000,
                            'startRow': start_row
                        }
                        
                        response = self.gsc_service.searchanalytics().query(
                            siteUrl=property_url, 
                            body=request
                        ).execute()
                        
                        rows = response.get('rows', [])
                        if not rows:
                            break
                            
                        all_rows.extend(rows)
                        
                        if len(rows) < 25000:
                            break
                            
                        start_row += 25000
                        
                        if start_row >= 100000:
                            break
                    
                    break
                    
                except Exception as e:
                    continue
            else:
                return pd.DataFrame()
            
            # Process data
            query_data = []
            for row in all_rows:
                query_data.append({
                    'query': row['keys'][0],
                    'clicks': row['clicks'],
                    'impressions': row['impressions'],
                    'ctr': round(row['ctr'] * 100, 2),
                    'position': round(row['position'], 1)
                })
            
            return pd.DataFrame(query_data)
            
        except Exception as e:
            print(f"❌ GSC query performance retrieval failed: {e}")
            return pd.DataFrame()
    
    def get_ga4_url_performance(self, property_id: str, start_date: str, end_date: str) -> pd.DataFrame:
        """Get GA4 sessions by URL"""
        print(f"📈 Fetching GA4 URL performance ({start_date} to {end_date})...")
        
        try:
            request = {
                'dateRanges': [{'startDate': start_date, 'endDate': end_date}],
                'dimensions': [{'name': 'pagePath'}],
                'metrics': [{'name': 'sessions'}],
                'limit': 100000
            }
            
            response = self.ga_service.properties().runReport(
                property=f'properties/{property_id}',
                body=request
            ).execute()
            
            url_data = []
            for row in response.get('rows', []):
                url_data.append({
                    'page_path': row['dimensionValues'][0]['value'],
                    'sessions': int(row['metricValues'][0]['value'])
                })
            
            return pd.DataFrame(url_data)
            
        except Exception as e:
            print(f"❌ GA4 URL performance retrieval failed: {e}")
            return pd.DataFrame()
    
    def create_excel_report(self, all_data: dict, output_file: str):
        """Create Excel file with all reports"""
        print("📄 Creating Excel report...")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Sheet 1: GSC Weekly Performance
            if not all_data['gsc_weekly'].empty:
                sheet_name = 'GSC Weekly Performance'
                all_data['gsc_weekly'].to_excel(writer, sheet_name=sheet_name, index=False)
                self.format_sheet(writer, sheet_name, 'GSC Weekly Performance')
            
            # Sheet 2: GA4 Weekly Performance
            if not all_data['ga4_weekly'].empty:
                sheet_name = 'GA4 Weekly Performance'
                all_data['ga4_weekly'].to_excel(writer, sheet_name=sheet_name, index=False)
                self.format_sheet(writer, sheet_name, 'GA4 Weekly Performance')
            
            # Sheet 3: 3-Month View
            if not all_data['three_month_view'].empty:
                sheet_name = '3-Month View'
                all_data['three_month_view'].to_excel(writer, sheet_name=sheet_name, index=False)
                self.format_sheet(writer, sheet_name, '3-Month Performance Comparison')
            
            # Sheet 4: Year over Year
            if not all_data['yoy_view'].empty:
                sheet_name = 'Year over Year'
                all_data['yoy_view'].to_excel(writer, sheet_name=sheet_name, index=False)
                self.format_sheet(writer, sheet_name, 'Year over Year Comparison')
            
            # Sheet 5: Top 10 Losing URLs (GSC)
            if not all_data['gsc_losing_urls'].empty:
                sheet_name = 'GSC Top 10 Losing URLs'
                all_data['gsc_losing_urls'].to_excel(writer, sheet_name=sheet_name, index=False)
                self.format_sheet(writer, sheet_name, 'Top 10 URLs with Biggest Click Drops')
            
            # Sheet 6: Top 25 Losing Queries
            if not all_data['gsc_losing_queries'].empty:
                sheet_name = 'GSC Top 25 Losing Queries'
                all_data['gsc_losing_queries'].to_excel(writer, sheet_name=sheet_name, index=False)
                self.format_sheet(writer, sheet_name, 'Top 25 Queries with Biggest Click Drops')
            
            # Sheet 7: GA4 Top 10 Losing URLs
            if not all_data['ga4_losing_urls'].empty:
                sheet_name = 'GA4 Top 10 Losing URLs'
                all_data['ga4_losing_urls'].to_excel(writer, sheet_name=sheet_name, index=False)
                self.format_sheet(writer, sheet_name, 'Top 10 URLs with Biggest Session Drops')
            
            # Sheet 8: GA4 Top 10 Winning URLs
            if not all_data['ga4_winning_urls'].empty:
                sheet_name = 'GA4 Top 10 Winning URLs'
                all_data['ga4_winning_urls'].to_excel(writer, sheet_name=sheet_name, index=False)
                self.format_sheet(writer, sheet_name, 'Top 10 URLs with Most Sessions')
        
        print(f"✅ Excel report saved to {output_file}")
    
    def format_sheet(self, writer, sheet_name: str, title: str):
        """Format Excel sheet with styling"""
        worksheet = writer.sheets[sheet_name]
        
        # Add title
        worksheet.insert_rows(1)
        worksheet['A1'] = title
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        worksheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        
        # Format headers
        for cell in worksheet[2]:
            if cell.value:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def generate_report(self, site_url: str, ga_property_id: str, start_date: str, end_date: str, output_file: str):
        """Main function to generate the All Hands report"""
        print("🚀 Starting All Hands report generation...")
        
        # Calculate date ranges
        date_ranges = self.get_date_ranges(start_date, end_date)
        print(f"📅 Date ranges calculated (period: {date_ranges['period_days']} days):")
        for period, dates in date_ranges.items():
            if period != 'period_days':
                start, end = dates
                print(f"   {period.capitalize()}: {start} to {end}")
        
        all_data = {}
        
        # 1. GSC Weekly Performance (current month)
        all_data['gsc_weekly'] = self.get_gsc_weekly_data(
            site_url, 
            date_ranges['current'][0], 
            date_ranges['current'][1]
        )
        
        # 2. GA4 Weekly Performance (current month)
        all_data['ga4_weekly'] = self.get_ga4_weekly_data(
            ga_property_id,
            date_ranges['current'][0],
            date_ranges['current'][1]
        )
        
        # 3. Three-Month Detailed View
        print("📊 Building 3-month detailed comparison...")
        
        # Get GSC data for all three months
        gsc_current = self.get_gsc_url_performance(site_url, date_ranges['current'][0], date_ranges['current'][1])
        gsc_previous = self.get_gsc_url_performance(site_url, date_ranges['previous'][0], date_ranges['previous'][1])
        gsc_prev2 = self.get_gsc_url_performance(site_url, date_ranges['previous2'][0], date_ranges['previous2'][1])
        
        # Get GA4 data for all three months
        ga4_current = self.get_ga4_monthly_data(ga_property_id, date_ranges['current'][0], date_ranges['current'][1])
        ga4_previous = self.get_ga4_monthly_data(ga_property_id, date_ranges['previous'][0], date_ranges['previous'][1])
        ga4_prev2 = self.get_ga4_monthly_data(ga_property_id, date_ranges['previous2'][0], date_ranges['previous2'][1])
        
        # Aggregate GSC data
        def aggregate_gsc_data(df):
            if df.empty:
                return {'clicks': 0, 'impressions': 0, 'avg_ctr': 0}
            return {
                'clicks': df['clicks'].sum(),
                'impressions': df['impressions'].sum(),
                'avg_ctr': df['ctr'].mean()
            }
        
        gsc_current_agg = aggregate_gsc_data(gsc_current)
        gsc_previous_agg = aggregate_gsc_data(gsc_previous)
        gsc_prev2_agg = aggregate_gsc_data(gsc_prev2)
        
        # Build 3-month comparison
        three_month_data = []
        
        # Current month
        three_month_data.append({
            'period': f"Current ({date_ranges['current'][0]} to {date_ranges['current'][1]})",
            'gsc_clicks': gsc_current_agg['clicks'],
            'gsc_impressions': gsc_current_agg['impressions'],
            'gsc_ctr': round(gsc_current_agg['avg_ctr'], 2),
            'ga4_total_sessions': ga4_current['total_sessions'],
            'ga4_organic_sessions': ga4_current['organic_sessions'],
            'ga4_paid_sessions': ga4_current['paid_sessions'],
            'ga4_direct_sessions': ga4_current['direct_sessions'],
            'gsc_clicks_change': self.calculate_percentage_change(gsc_previous_agg['clicks'], gsc_current_agg['clicks']),
            'gsc_impressions_change': self.calculate_percentage_change(gsc_previous_agg['impressions'], gsc_current_agg['impressions']),
            'gsc_ctr_change': self.calculate_percentage_change(gsc_previous_agg['avg_ctr'], gsc_current_agg['avg_ctr']),
            'gsc_ctr_change_points': round(gsc_current_agg['avg_ctr'] - gsc_previous_agg['avg_ctr'], 2),
            'ga4_total_sessions_change': self.calculate_percentage_change(ga4_previous['total_sessions'], ga4_current['total_sessions']),
            'ga4_organic_sessions_change': self.calculate_percentage_change(ga4_previous['organic_sessions'], ga4_current['organic_sessions']),
            'ga4_paid_sessions_change': self.calculate_percentage_change(ga4_previous['paid_sessions'], ga4_current['paid_sessions']),
            'ga4_direct_sessions_change': self.calculate_percentage_change(ga4_previous['direct_sessions'], ga4_current['direct_sessions'])
        })
        
        # Previous month
        three_month_data.append({
            'period': f"Previous ({date_ranges['previous'][0]} to {date_ranges['previous'][1]})",
            'gsc_clicks': gsc_previous_agg['clicks'],
            'gsc_impressions': gsc_previous_agg['impressions'],
            'gsc_ctr': round(gsc_previous_agg['avg_ctr'], 2),
            'ga4_total_sessions': ga4_previous['total_sessions'],
            'ga4_organic_sessions': ga4_previous['organic_sessions'],
            'ga4_paid_sessions': ga4_previous['paid_sessions'],
            'ga4_direct_sessions': ga4_previous['direct_sessions'],
            'gsc_clicks_change': self.calculate_percentage_change(gsc_prev2_agg['clicks'], gsc_previous_agg['clicks']),
            'gsc_impressions_change': self.calculate_percentage_change(gsc_prev2_agg['impressions'], gsc_previous_agg['impressions']),
            'gsc_ctr_change': self.calculate_percentage_change(gsc_prev2_agg['avg_ctr'], gsc_previous_agg['avg_ctr']),
            'gsc_ctr_change_points': round(gsc_previous_agg['avg_ctr'] - gsc_prev2_agg['avg_ctr'], 2),
            'ga4_total_sessions_change': self.calculate_percentage_change(ga4_prev2['total_sessions'], ga4_previous['total_sessions']),
            'ga4_organic_sessions_change': self.calculate_percentage_change(ga4_prev2['organic_sessions'], ga4_previous['organic_sessions']),
            'ga4_paid_sessions_change': self.calculate_percentage_change(ga4_prev2['paid_sessions'], ga4_previous['paid_sessions']),
            'ga4_direct_sessions_change': self.calculate_percentage_change(ga4_prev2['direct_sessions'], ga4_previous['direct_sessions'])
        })
        
        # Month -2
        three_month_data.append({
            'period': f"Month-2 ({date_ranges['previous2'][0]} to {date_ranges['previous2'][1]})",
            'gsc_clicks': gsc_prev2_agg['clicks'],
            'gsc_impressions': gsc_prev2_agg['impressions'],
            'gsc_ctr': round(gsc_prev2_agg['avg_ctr'], 2),
            'ga4_total_sessions': ga4_prev2['total_sessions'],
            'ga4_organic_sessions': ga4_prev2['organic_sessions'],
            'ga4_paid_sessions': ga4_prev2['paid_sessions'],
            'ga4_direct_sessions': ga4_prev2['direct_sessions'],
            'gsc_clicks_change': 0,  # No previous period for oldest month
            'gsc_impressions_change': 0,
            'gsc_ctr_change': 0,
            'gsc_ctr_change_points': 0,
            'ga4_total_sessions_change': 0,
            'ga4_organic_sessions_change': 0,
            'ga4_paid_sessions_change': 0,
            'ga4_direct_sessions_change': 0
        })
        
        all_data['three_month_view'] = pd.DataFrame(three_month_data)
        
        # 4. Year over Year Detailed View
        print("📊 Building year-over-year comparison...")
        
        gsc_yoy = self.get_gsc_url_performance(site_url, date_ranges['yoy'][0], date_ranges['yoy'][1])
        ga4_yoy = self.get_ga4_monthly_data(ga_property_id, date_ranges['yoy'][0], date_ranges['yoy'][1])
        
        gsc_yoy_agg = aggregate_gsc_data(gsc_yoy)
        
        yoy_data = []
        
        # Current year data
        yoy_data.append({
            'period': f"Current Year ({date_ranges['current'][0]} to {date_ranges['current'][1]})",
            'gsc_clicks': gsc_current_agg['clicks'],
            'gsc_impressions': gsc_current_agg['impressions'],
            'gsc_ctr': round(gsc_current_agg['avg_ctr'], 2),
            'ga4_total_sessions': ga4_current['total_sessions'],
            'ga4_organic_sessions': ga4_current['organic_sessions'],
            'ga4_paid_sessions': ga4_current['paid_sessions'],
            'ga4_direct_sessions': ga4_current['direct_sessions'],
            'gsc_clicks_yoy_change': self.calculate_percentage_change(gsc_yoy_agg['clicks'], gsc_current_agg['clicks']),
            'gsc_impressions_yoy_change': self.calculate_percentage_change(gsc_yoy_agg['impressions'], gsc_current_agg['impressions']),
            'gsc_ctr_yoy_change': self.calculate_percentage_change(gsc_yoy_agg['avg_ctr'], gsc_current_agg['avg_ctr']),
            'gsc_ctr_yoy_change_points': round(gsc_current_agg['avg_ctr'] - gsc_yoy_agg['avg_ctr'], 2),
            'ga4_total_sessions_yoy_change': self.calculate_percentage_change(ga4_yoy['total_sessions'], ga4_current['total_sessions']),
            'ga4_organic_sessions_yoy_change': self.calculate_percentage_change(ga4_yoy['organic_sessions'], ga4_current['organic_sessions']),
            'ga4_paid_sessions_yoy_change': self.calculate_percentage_change(ga4_yoy['paid_sessions'], ga4_current['paid_sessions']),
            'ga4_direct_sessions_yoy_change': self.calculate_percentage_change(ga4_yoy['direct_sessions'], ga4_current['direct_sessions'])
        })
        
        # Previous year data
        yoy_data.append({
            'period': f"Previous Year ({date_ranges['yoy'][0]} to {date_ranges['yoy'][1]})",
            'gsc_clicks': gsc_yoy_agg['clicks'],
            'gsc_impressions': gsc_yoy_agg['impressions'],
            'gsc_ctr': round(gsc_yoy_agg['avg_ctr'], 2),
            'ga4_total_sessions': ga4_yoy['total_sessions'],
            'ga4_organic_sessions': ga4_yoy['organic_sessions'],
            'ga4_paid_sessions': ga4_yoy['paid_sessions'],
            'ga4_direct_sessions': ga4_yoy['direct_sessions'],
            'gsc_clicks_yoy_change': 0,  # Reference point
            'gsc_impressions_yoy_change': 0,
            'gsc_ctr_yoy_change': 0,
            'gsc_ctr_yoy_change_points': 0,
            'ga4_total_sessions_yoy_change': 0,
            'ga4_organic_sessions_yoy_change': 0,
            'ga4_paid_sessions_yoy_change': 0,
            'ga4_direct_sessions_yoy_change': 0
        })
        
        all_data['yoy_view'] = pd.DataFrame(yoy_data)
        
        # 5. Top 10 Losing URLs (GSC)
        if not gsc_current.empty and not gsc_previous.empty:
            # Merge current and previous data
            gsc_comparison = gsc_current.merge(gsc_previous, on='url', how='outer', suffixes=('_current', '_previous'))
            gsc_comparison = gsc_comparison.fillna(0)
            
            # Calculate changes with percentage points for CTR
            gsc_comparison['clicks_change'] = gsc_comparison['clicks_current'] - gsc_comparison['clicks_previous']
            gsc_comparison['impressions_mom_change'] = gsc_comparison.apply(
                lambda row: self.calculate_percentage_change(row['impressions_previous'], row['impressions_current']), axis=1
            )
            gsc_comparison['ctr_mom_change'] = gsc_comparison.apply(
                lambda row: self.calculate_percentage_change(row['ctr_previous'], row['ctr_current']), axis=1
            )
            gsc_comparison['ctr_change_points'] = gsc_comparison['ctr_current'] - gsc_comparison['ctr_previous']
            gsc_comparison['ctr_change_points'] = gsc_comparison['ctr_change_points'].round(2)
            
            # Get top 10 losers
            gsc_losing = gsc_comparison.nsmallest(10, 'clicks_change')[
                ['url', 'clicks_current', 'clicks_previous', 'clicks_change', 'impressions_mom_change', 'ctr_mom_change', 'ctr_change_points']
            ]
            gsc_losing = gsc_losing.rename(columns={
                'clicks_current': 'current_clicks',
                'clicks_previous': 'previous_clicks'
            })
            
            all_data['gsc_losing_urls'] = gsc_losing
        else:
            all_data['gsc_losing_urls'] = pd.DataFrame()
        
        # 6. Top 25 Losing Queries (GSC)
        gsc_queries_current = self.get_gsc_query_performance(site_url, date_ranges['current'][0], date_ranges['current'][1])
        gsc_queries_previous = self.get_gsc_query_performance(site_url, date_ranges['previous'][0], date_ranges['previous'][1])
        
        if not gsc_queries_current.empty and not gsc_queries_previous.empty:
            # Merge current and previous data
            queries_comparison = gsc_queries_current.merge(gsc_queries_previous, on='query', how='outer', suffixes=('_current', '_previous'))
            queries_comparison = queries_comparison.fillna(0)
            
            # Calculate changes with percentage points for CTR
            queries_comparison['clicks_change'] = queries_comparison['clicks_current'] - queries_comparison['clicks_previous']
            queries_comparison['impressions_mom_change'] = queries_comparison.apply(
                lambda row: self.calculate_percentage_change(row['impressions_previous'], row['impressions_current']), axis=1
            )
            queries_comparison['ctr_mom_change'] = queries_comparison.apply(
                lambda row: self.calculate_percentage_change(row['ctr_previous'], row['ctr_current']), axis=1
            )
            queries_comparison['ctr_change_points'] = queries_comparison['ctr_current'] - queries_comparison['ctr_previous']
            queries_comparison['ctr_change_points'] = queries_comparison['ctr_change_points'].round(2)
            
            # Get top 25 losers
            queries_losing = queries_comparison.nsmallest(25, 'clicks_change')[
                ['query', 'clicks_current', 'clicks_previous', 'clicks_change', 'impressions_mom_change', 'ctr_mom_change', 'ctr_change_points']
            ]
            queries_losing = queries_losing.rename(columns={
                'clicks_current': 'current_clicks',
                'clicks_previous': 'previous_clicks'
            })
            
            all_data['gsc_losing_queries'] = queries_losing
        else:
            all_data['gsc_losing_queries'] = pd.DataFrame()
        
        # 7. GA4 Top 10 Losing URLs
        ga4_urls_current = self.get_ga4_url_performance(ga_property_id, date_ranges['current'][0], date_ranges['current'][1])
        ga4_urls_previous = self.get_ga4_url_performance(ga_property_id, date_ranges['previous'][0], date_ranges['previous'][1])
        
        if not ga4_urls_current.empty and not ga4_urls_previous.empty:
            # Merge current and previous data
            ga4_comparison = ga4_urls_current.merge(ga4_urls_previous, on='page_path', how='outer', suffixes=('_current', '_previous'))
            ga4_comparison = ga4_comparison.fillna(0)
            
            # Calculate changes
            ga4_comparison['sessions_change'] = ga4_comparison['sessions_current'] - ga4_comparison['sessions_previous']
            ga4_comparison['sessions_mom_change'] = ga4_comparison.apply(
                lambda row: self.calculate_percentage_change(row['sessions_previous'], row['sessions_current']), axis=1
            )
            
            # Get top 10 losers
            ga4_losing = ga4_comparison.nsmallest(10, 'sessions_change')[
                ['page_path', 'sessions_current', 'sessions_previous', 'sessions_change', 'sessions_mom_change']
            ]
            ga4_losing = ga4_losing.rename(columns={
                'sessions_current': 'current_sessions',
                'sessions_previous': 'previous_sessions'
            })
            
            all_data['ga4_losing_urls'] = ga4_losing
        else:
            all_data['ga4_losing_urls'] = pd.DataFrame()
        
        # 8. GA4 Top 10 Winning URLs (by absolute sessions)
        if not ga4_urls_current.empty:
            ga4_winning = ga4_urls_current.nlargest(10, 'sessions').copy()
            
            # Add MoM change if previous data exists
            if not ga4_urls_previous.empty:
                ga4_winning = ga4_winning.merge(ga4_urls_previous, on='page_path', how='left', suffixes=('_current', '_previous'))
                ga4_winning['sessions_previous'] = ga4_winning['sessions_previous'].fillna(0)
                ga4_winning['sessions_mom_change'] = ga4_winning.apply(
                    lambda row: self.calculate_percentage_change(row['sessions_previous'], row['sessions_current']), axis=1
                )
                ga4_winning = ga4_winning[['page_path', 'sessions_current', 'sessions_previous', 'sessions_mom_change']]
                ga4_winning = ga4_winning.rename(columns={
                    'sessions_current': 'current_sessions',
                    'sessions_previous': 'previous_sessions'
                })
            else:
                ga4_winning = ga4_winning[['page_path', 'sessions']]
                ga4_winning = ga4_winning.rename(columns={'sessions': 'current_sessions'})
                ga4_winning['previous_sessions'] = 0
                ga4_winning['sessions_mom_change'] = 0
            
            all_data['ga4_winning_urls'] = ga4_winning
        else:
            all_data['ga4_winning_urls'] = pd.DataFrame()
        
        # Create Excel report
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        self.create_excel_report(all_data, output_file)
        
        # Print summary
        print(f"\n📊 All Hands Report Summary:")
        print(f"   Period analyzed: {date_ranges['current'][0]} to {date_ranges['current'][1]} ({date_ranges['period_days']} days)")
        print(f"   GSC Weekly data points: {len(all_data['gsc_weekly'])}")
        print(f"   GA4 Weekly data points: {len(all_data['ga4_weekly'])}")
        if not all_data['three_month_view'].empty:
            current_row = all_data['three_month_view'].iloc[0]
            print(f"   GA4 total sessions (current): {current_row['ga4_total_sessions']:,}")
            print(f"   GA4 organic sessions (current): {current_row['ga4_organic_sessions']:,}")
            print(f"   GSC total clicks (current): {current_row['gsc_clicks']:,}")
        print(f"   Top losing URLs identified: {len(all_data['gsc_losing_urls'])} (GSC), {len(all_data['ga4_losing_urls'])} (GA4)")
        print(f"   Top winning URLs identified: {len(all_data['ga4_winning_urls'])} (GA4)")
        
        return all_data

def main():
    print("🚀 All Hands Report Generator Starting...")
    print("=" * 50)
    
    parser = argparse.ArgumentParser(description='All Hands Report Generator')
    parser.add_argument('--site-url', required=True, help='GSC site URL (e.g., https://www.example.com/)')
    parser.add_argument('--ga-property-id', required=True, help='GA4 Property ID (numbers only)')
    parser.add_argument('--credentials', required=True, help='Path to service account credentials JSON')
    parser.add_argument('--start-date', required=True, help='Start date for analysis (YYYY-MM-DD)')
    parser.add_argument('--end-date', required=True, help='End date for analysis (YYYY-MM-DD)')
    parser.add_argument('--output', default='AllHandsReports/all_hands_report.xlsx', help='Output Excel file path')
    
    try:
        args = parser.parse_args()
        print(f"✓ Arguments parsed successfully")
        print(f"  Site URL: {args.site_url}")
        print(f"  GA Property ID: {args.ga_property_id}")
        print(f"  Credentials: {args.credentials}")
        print(f"  Start Date: {args.start_date}")
        print(f"  End Date: {args.end_date}")
        print(f"  Output: {args.output}")
        print()
        
    except Exception as e:
        print(f"❌ Failed to parse arguments: {e}")
        return
    
    # Check if credentials file exists
    if not os.path.exists(args.credentials):
        print(f"❌ Credentials file not found: {args.credentials}")
        return
    else:
        print(f"✓ Credentials file found")
    
    # Validate date formats
    try:
        start_dt = datetime.strptime(args.start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(args.end_date, '%Y-%m-%d')
        
        if start_dt >= end_dt:
            print("❌ Start date must be before end date")
            return
            
        period_days = (end_dt - start_dt).days + 1
        print(f"✓ Date format valid ({period_days} days period)")
        
    except ValueError:
        print("❌ Invalid date format. Please use YYYY-MM-DD for both dates")
        return
    
    # Check required packages
    try:
        import pandas as pd
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
        import openpyxl
        print("✓ All required packages available")
    except ImportError as e:
        print(f"❌ Missing required package: {e}")
        print("Please install: pip install pandas google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client openpyxl")
        return
    
    print("\n🔧 Initializing APIs...")
    
    # Initialize tool
    try:
        tool = AllHandsReportGenerator(credentials_path=args.credentials)
        print("✓ Tool initialized successfully")
    except Exception as e:
        print(f"❌ Failed to initialize tool: {e}")
        print(f"   Error details: {str(e)}")
        return
    
    # Generate report
    print(f"\n📊 Starting report generation...")
    try:
        tool.generate_report(
            site_url=args.site_url,
            ga_property_id=args.ga_property_id,
            start_date=args.start_date,
            end_date=args.end_date,
            output_file=args.output
        )
        print(f"\n✅ All Hands report generated successfully!")
        print(f"📁 Report saved to: {args.output}")
        
    except Exception as e:
        print(f"❌ Report generation failed: {e}")
        print(f"   Error type: {type(e).__name__}")
        print(f"   Error details: {str(e)}")
        import traceback
        print(f"   Traceback:")
        traceback.print_exc()

if __name__ == "__main__":
    main()
